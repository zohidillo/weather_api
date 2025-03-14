from openpyxl.styles import (
    Font, Alignment, Border, Side
)

column_widths = {
    "A": 30,
    "B": 30,
    "C": 30,
    "D": 30,
    "E": 30,
    "F": 30,
    "G": 30,
    "H": 30,
    "I": 30,
    "J": 30,
}


def set_title(ws, title):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    title_cell = ws['A1']
    title_cell.value = title


def excel_style(ws):
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    row_height = 35
    for row in ws.iter_rows():
        for cell in row:
            ws.row_dimensions[cell.row].height = row_height

    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    header_font = Font(bold=True, size=14, name="Helvetica")
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # set up title
    ws.row_dimensions[1].height = 100
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=28, name="Helvetica")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
